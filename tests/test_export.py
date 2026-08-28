"""Export: PDF splitting quality, output layout, naming and collisions."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pymupdf
import pytest

from app.models.candidate import Candidate
from app.models.document import DocumentGroup
from app.models.enums import FileStatus
from app.models.source_file import SourceFileAnalysis
from app.services.excel_service import write_document_index
from app.services.export_service import (
    ExportService,
    batch_folder_name,
    create_batch_directory,
)
from app.services.pdf_service import (
    PdfError,
    contiguous_runs,
    open_pdf,
    split_pdf,
)
from scripts import sample_data


@pytest.fixture
def packet(samples_dir: Path) -> Path:
    return samples_dir / sample_data.sample_a().filename


class TestContiguousRuns:
    def test_single_run(self) -> None:
        assert contiguous_runs([0, 1, 2]) == [(0, 2)]

    def test_split_runs(self) -> None:
        assert contiguous_runs([0, 1, 4, 5, 9]) == [(0, 1), (4, 5), (9, 9)]

    def test_unsorted_and_duplicated_input(self) -> None:
        assert contiguous_runs([5, 4, 4, 0]) == [(0, 0), (4, 5)]

    def test_empty(self) -> None:
        assert contiguous_runs([]) == []


class TestSplitPdf:
    def test_page_count_is_preserved(self, packet: Path, tmp_path: Path) -> None:
        output = split_pdf(packet, [4, 5, 6], tmp_path / "resume.pdf")
        with open_pdf(output) as document:
            assert document.page_count == 3

    def test_page_order_is_preserved(self, packet: Path, tmp_path: Path) -> None:
        output = split_pdf(packet, [4, 5, 6], tmp_path / "resume.pdf")
        with open_pdf(output) as extracted, open_pdf(packet) as source:
            for offset, page_index in enumerate([4, 5, 6]):
                assert extracted.load_page(offset).get_text("text").strip() == (
                    source.load_page(page_index).get_text("text").strip()
                )

    def test_searchable_text_survives(self, packet: Path, tmp_path: Path) -> None:
        output = split_pdf(packet, [7], tmp_path / "letter.pdf")
        with open_pdf(output) as document:
            text = document.load_page(0).get_text("text")
        assert "Dear Ms. Ortiz" in text
        assert "Sincerely" in text

    def test_page_dimensions_and_rotation_are_preserved(
        self, packet: Path, tmp_path: Path
    ) -> None:
        output = split_pdf(packet, [0, 1], tmp_path / "report.pdf")
        with open_pdf(output) as extracted, open_pdf(packet) as source:
            for offset in range(2):
                assert extracted.load_page(offset).rect == source.load_page(offset).rect
                assert extracted.load_page(offset).rotation == source.load_page(offset).rotation

    def test_pages_are_not_rasterised(self, packet: Path, tmp_path: Path) -> None:
        """A split page must remain real text, not a picture of text."""
        output = split_pdf(packet, [4], tmp_path / "page.pdf")
        with open_pdf(output) as document:
            page = document.load_page(0)
            assert len(page.get_text("text").strip()) > 100
            assert not page.get_images(full=True), "page was converted to an image"

    def test_single_page_export(self, packet: Path, tmp_path: Path) -> None:
        output = split_pdf(packet, [7], tmp_path / "one.pdf")
        with open_pdf(output) as document:
            assert document.page_count == 1

    def test_non_contiguous_pages(self, packet: Path, tmp_path: Path) -> None:
        output = split_pdf(packet, [0, 9], tmp_path / "ends.pdf")
        with open_pdf(output) as document:
            assert document.page_count == 2

    def test_out_of_range_pages_are_rejected(self, packet: Path, tmp_path: Path) -> None:
        with pytest.raises(PdfError):
            split_pdf(packet, [99], tmp_path / "bad.pdf")

    def test_empty_selection_is_rejected(self, packet: Path, tmp_path: Path) -> None:
        with pytest.raises(PdfError):
            split_pdf(packet, [], tmp_path / "bad.pdf")

    def test_no_partial_file_is_left_behind_on_failure(
        self, packet: Path, tmp_path: Path
    ) -> None:
        destination = tmp_path / "bad.pdf"
        with pytest.raises(PdfError):
            split_pdf(packet, [99], destination)
        assert not destination.exists()
        assert not list(tmp_path.glob("*.part"))


def build_analysis(source: Path, groups: list[DocumentGroup]) -> SourceFileAnalysis:
    analysis = SourceFileAnalysis(path=source, status=FileStatus.READY)
    analysis.groups = groups
    return analysis


class TestExportService:
    def test_creates_candidate_folders(self, pipeline, packet: Path, tmp_path: Path) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService().export([analysis], tmp_path)

        assert result.document_count == 4
        folder = tmp_path / "Benjamin Perez"
        assert folder.is_dir()
        names = sorted(p.name for p in folder.glob("*.pdf"))
        assert names == [
            "Benjamin_Perez_Application_Report.pdf",
            # Both output modes are on by default, so the candidate also gets
            # their whole packet as one file.
            "Benjamin_Perez_Complete_Packet.pdf",
            "Benjamin_Perez_Cover_Letter.pdf",
            "Benjamin_Perez_References.pdf",
            "Benjamin_Perez_Resume.pdf",
        ]

    def test_separate_documents_only(self, pipeline, packet: Path, tmp_path: Path) -> None:
        analysis = pipeline.analyze_file(packet)
        ExportService(export_combined_packets=False).export([analysis], tmp_path)

        folder = tmp_path / "Benjamin Perez"
        names = sorted(p.name for p in folder.glob("*.pdf"))
        assert names == [
            "Benjamin_Perez_Application_Report.pdf",
            "Benjamin_Perez_Cover_Letter.pdf",
            "Benjamin_Perez_References.pdf",
            "Benjamin_Perez_Resume.pdf",
        ]

    def test_combined_packet_only(self, pipeline, packet: Path, tmp_path: Path) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService(export_separate_documents=False).export([analysis], tmp_path)

        assert result.document_count == 0
        assert result.packet_count == 1
        folder = tmp_path / "Benjamin Perez"
        assert sorted(p.name for p in folder.glob("*.pdf")) == [
            "Benjamin_Perez_Complete_Packet.pdf"
        ]

    def test_exported_documents_have_the_right_page_counts(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        ExportService().export([analysis], tmp_path)
        folder = tmp_path / "Benjamin Perez"

        expected = {
            "Benjamin_Perez_Application_Report.pdf": 4,
            "Benjamin_Perez_Resume.pdf": 3,
            "Benjamin_Perez_Cover_Letter.pdf": 1,
            "Benjamin_Perez_References.pdf": 2,
        }
        for name, pages in expected.items():
            with open_pdf(folder / name) as document:
                assert document.page_count == pages, name

    def test_flat_output_when_folders_disabled(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        ExportService(folder_per_candidate=False).export([analysis], tmp_path)
        # Four documents plus the candidate's combined packet, no subfolders.
        assert len(list(tmp_path.glob("*.pdf"))) == 5
        assert not list(tmp_path.glob("*/"))

    def test_existing_files_are_never_overwritten(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        service = ExportService()
        first = pipeline.analyze_file(packet)
        service.export([first], tmp_path)
        second = pipeline.analyze_file(packet)
        service.export([second], tmp_path)

        folder = tmp_path / "Benjamin Perez"
        assert (folder / "Benjamin_Perez_Resume.pdf").exists()
        assert (folder / "Benjamin_Perez_Resume_2.pdf").exists()
        # The combined packet must not overwrite the first run's copy either.
        assert (folder / "Benjamin_Perez_Complete_Packet.pdf").exists()
        assert (folder / "Benjamin_Perez_Complete_Packet_2.pdf").exists()
        assert len(list(folder.glob("*.pdf"))) == 10

    def test_excluded_documents_are_not_written(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        analysis.groups[0].excluded = True
        result = ExportService().export([analysis], tmp_path)
        assert result.document_count == 3
        assert not (tmp_path / "Benjamin Perez" / "Benjamin_Perez_Application_Report.pdf").exists()

    def test_unknown_candidate_goes_to_unknown_folder_with_sequence(
        self, packet: Path, tmp_path: Path
    ) -> None:
        group = DocumentGroup(
            source_pdf=str(packet), page_indexes=[0], document_type="Resume"
        )
        group.candidate = Candidate()
        analysis = build_analysis(packet, [group])

        ExportService().export([analysis], tmp_path)
        folder = tmp_path / "Unknown"
        assert folder.is_dir()
        assert [p.name for p in folder.glob("*.pdf")] == ["Unknown_Resume_001.pdf"]

    def test_custom_filename_template(self, pipeline, packet: Path, tmp_path: Path) -> None:
        service = ExportService(filename_template="{document_type}-{candidate}")
        analysis = pipeline.analyze_file(packet)
        service.export([analysis], tmp_path)
        names = sorted(p.name for p in (tmp_path / "Benjamin Perez").glob("*.pdf"))
        assert "Resume-Benjamin_Perez.pdf" in names

    def test_separator_pages_excluded_from_output_are_not_written(
        self, profile, thresholds, samples_dir: Path, tmp_path: Path
    ) -> None:
        from app.models.enums import SeparatorPolicy
        from tests.helpers import build_pipeline

        pipeline = build_pipeline(profile, thresholds, separator_policy=SeparatorPolicy.EXCLUDE)
        analysis = pipeline.analyze_file(samples_dir / sample_data.sample_f().filename)
        ExportService().export([analysis], tmp_path)

        resume = next(p for p in tmp_path.rglob("*Resume*.pdf"))
        with open_pdf(resume) as document:
            assert document.page_count == 2  # separator dropped

    def test_a_failing_file_does_not_stop_the_batch(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        good = pipeline.analyze_file(packet)
        broken = build_analysis(
            tmp_path / "missing.pdf",
            [DocumentGroup(source_pdf=str(tmp_path / "missing.pdf"), page_indexes=[0])],
        )

        result = ExportService().export([broken, good], tmp_path / "out")
        assert result.has_errors
        assert result.document_count == 4

    def test_missing_output_directory_is_created(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        destination = tmp_path / "deeply" / "nested" / "output"
        analysis = pipeline.analyze_file(packet)
        result = ExportService().export([analysis], destination)
        assert destination.is_dir()
        assert result.document_count == 4


def read_index(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    """The index as ``(headers, rows-as-dicts)``.

    Addressing columns by name rather than position: a positional test breaks
    the moment a column is inserted, which says nothing about whether the
    spreadsheet is correct.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    rows = [
        dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=2, values_only=True)
    ]
    workbook.close()
    return headers, rows


class TestExcelIndex:
    def test_index_is_written_with_a_row_per_document(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        from openpyxl import load_workbook

        analysis = pipeline.analyze_file(packet)
        result = ExportService().export([analysis], tmp_path)
        path = write_document_index(result.exported, tmp_path, packets=analysis.packets)

        assert path is not None and path.exists()
        workbook = load_workbook(path)
        sheet = workbook.active
        assert sheet.max_row == 5  # header + 4 documents
        assert sheet.freeze_panes == "A2"
        assert sheet.auto_filter.ref is not None
        workbook.close()

    def test_index_contains_expected_values(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService().export([analysis], tmp_path)
        path = write_document_index(result.exported, tmp_path, packets=analysis.packets)

        _headers, rows = read_index(path)
        assert {row["Candidate"] for row in rows} == {"Benjamin Perez"}
        assert {row["Document Type"] for row in rows} == {
            "Application Report",
            "Resume",
            "Cover Letter",
            "References",
        }

    def test_index_records_the_candidate_packet(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService().export([analysis], tmp_path)
        path = write_document_index(result.exported, tmp_path, packets=analysis.packets)

        _headers, rows = read_index(path)
        packet_ids = {row["Candidate Packet"] for row in rows}
        assert len(packet_ids) == 1, "one candidate's documents span several packets"
        assert all(row["Candidate Confidence"] for row in rows)
        assert all(
            str(row["Combined Packet File"]).endswith("_Complete_Packet.pdf") for row in rows
        )

    def test_confidence_columns_are_formatted_as_percentages(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        """A column insert must not leave the percentage format on the wrong one."""
        from openpyxl import load_workbook

        analysis = pipeline.analyze_file(packet)
        result = ExportService().export([analysis], tmp_path)
        path = write_document_index(result.exported, tmp_path, packets=analysis.packets)

        workbook = load_workbook(path)
        sheet = workbook.active
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=index)
            if str(header).endswith("Confidence"):
                assert cell.number_format == "0.0%", header
            else:
                assert cell.number_format != "0.0%", header
        workbook.close()

    def test_no_index_for_an_empty_export(self, tmp_path: Path) -> None:
        assert write_document_index([], tmp_path) is None


class TestCombinedPacket:
    """One PDF per candidate, in the configured order, at original quality."""

    @pytest.fixture
    def out_of_order(self, tmp_path: Path) -> Path:
        """A source PDF whose documents are in the wrong order for a packet.

        Resume first, then the application report, then the cover letter --
        the packet must reorder them without touching page order inside each.
        """
        from scripts.mixed_batch import (
            APPLICANTS,
            MixedBatchBuilder,
            _application_report,
            _cover_letter,
            _resume,
        )
        from app.profiles.recruiting import APPLICATION_REPORT, COVER_LETTER, RESUME

        applicant = APPLICANTS[0]
        builder = MixedBatchBuilder("out_of_order.pdf")
        builder.start_candidate(applicant.name)
        builder.add_document(RESUME, _resume(applicant, 2))
        builder.add_document(APPLICATION_REPORT, _application_report(applicant, 3))
        builder.add_document(COVER_LETTER, _cover_letter(applicant))
        return builder.build().write(tmp_path)

    def test_the_packet_follows_the_configured_order(
        self, pipeline, out_of_order: Path, tmp_path: Path
    ) -> None:
        from app.profiles.recruiting import DEFAULT_PACKET_ORDER

        analysis = pipeline.analyze_file(out_of_order)
        packet = analysis.identified_packets[0]
        ordered = [d.document_type for d in packet.ordered_documents(DEFAULT_PACKET_ORDER)]

        assert ordered == ["Application Report", "Resume", "Cover Letter"], (
            "the packet kept source order instead of the configured order"
        )

    def test_page_order_inside_a_document_is_preserved(
        self, pipeline, out_of_order: Path
    ) -> None:
        from app.profiles.recruiting import DEFAULT_PACKET_ORDER

        analysis = pipeline.analyze_file(out_of_order)
        packet = analysis.identified_packets[0]
        for document in packet.ordered_documents(DEFAULT_PACKET_ORDER):
            pages = document.export_page_indexes
            assert pages == sorted(pages), "pages were reordered inside a document"

    def test_a_missing_type_is_simply_skipped(self, pipeline, tmp_path: Path) -> None:
        from scripts.mixed_batch import APPLICANTS, MixedBatchBuilder, _cover_letter, _resume
        from app.profiles.recruiting import COVER_LETTER, DEFAULT_PACKET_ORDER, RESUME

        applicant = APPLICANTS[2]
        builder = MixedBatchBuilder("no_report.pdf")
        builder.start_candidate(applicant.name)
        builder.add_document(RESUME, _resume(applicant, 2))
        builder.add_document(COVER_LETTER, _cover_letter(applicant))
        source = builder.build().write(tmp_path)

        analysis = pipeline.analyze_file(source)
        packet = analysis.identified_packets[0]
        types = [d.document_type for d in packet.ordered_documents(DEFAULT_PACKET_ORDER)]
        assert types == ["Resume", "Cover Letter"]

    def test_the_combined_pdf_is_written_in_that_order(
        self, pipeline, out_of_order: Path, tmp_path: Path
    ) -> None:
        """The order on disk, not just in memory."""
        from app.profiles.recruiting import DEFAULT_PACKET_ORDER

        analysis = pipeline.analyze_file(out_of_order)
        destination = tmp_path / "out"
        ExportService(packet_order=DEFAULT_PACKET_ORDER).export([analysis], destination)

        combined = next(destination.rglob("*_Complete_Packet.pdf"))
        with open_pdf(combined) as document:
            first_page = document.load_page(0).get_text("text")
        assert "APPLICATION REPORT" in first_page.upper(), (
            "the combined packet does not open with the application report"
        )

    def test_the_combined_pdf_copies_pages_rather_than_rasterising(
        self, pipeline, out_of_order: Path, tmp_path: Path
    ) -> None:
        """Same quality guarantee as splitting: no re-encoding, text intact."""
        from app.profiles.recruiting import DEFAULT_PACKET_ORDER

        analysis = pipeline.analyze_file(out_of_order)
        destination = tmp_path / "out"
        ExportService(packet_order=DEFAULT_PACKET_ORDER).export([analysis], destination)
        combined = next(destination.rglob("*_Complete_Packet.pdf"))

        with open_pdf(out_of_order) as source, open_pdf(combined) as result:
            assert result.page_count == source.page_count
            source_size = source.load_page(0).rect
            for index in range(result.page_count):
                page = result.load_page(index)
                assert page.get_text("text").strip(), f"page {index + 1} lost its text"
                assert not page.get_images(), "an unexpected image appeared"
            assert result.load_page(0).rect.width == source_size.width
            assert result.load_page(0).rect.height == source_size.height

    def test_excluded_documents_stay_out_of_the_packet(
        self, pipeline, out_of_order: Path, tmp_path: Path
    ) -> None:
        from app.profiles.recruiting import DEFAULT_PACKET_ORDER

        analysis = pipeline.analyze_file(out_of_order)
        dropped = analysis.groups[0]
        dropped.excluded = True

        destination = tmp_path / "out"
        ExportService(packet_order=DEFAULT_PACKET_ORDER).export([analysis], destination)
        combined = next(destination.rglob("*_Complete_Packet.pdf"))

        expected_pages = sum(
            len(d.export_page_indexes)
            for d in analysis.identified_packets[0].ordered_documents(DEFAULT_PACKET_ORDER)
        )
        with open_pdf(combined) as document:
            assert document.page_count == expected_pages


class TestDocumentTypeFilter:
    """Saving only some document types -- resumes only, say."""

    def test_everything_is_saved_by_default(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService().export([analysis], tmp_path)
        assert result.document_count == 4

    def test_only_the_chosen_type_is_written(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService(document_types=["Resume"]).export([analysis], tmp_path)

        assert [d.group.document_type for d in result.exported] == ["Resume"]
        written = sorted(p.name for p in tmp_path.rglob("*.pdf"))
        assert "Benjamin_Perez_Resume.pdf" in written
        assert not any("Cover_Letter" in name for name in written)

    def test_several_types_can_be_chosen(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService(
            document_types=["Resume", "Cover Letter"], export_combined_packets=False
        ).export([analysis], tmp_path)

        assert sorted(d.group.document_type for d in result.exported) == [
            "Cover Letter",
            "Resume",
        ]

    def test_the_combined_packet_honours_the_filter(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        """A packet of "resumes only" must not quietly contain everything."""
        from app.profiles.recruiting import DEFAULT_PACKET_ORDER

        analysis = pipeline.analyze_file(packet)
        ExportService(
            document_types=["Resume"],
            export_separate_documents=False,
            packet_order=DEFAULT_PACKET_ORDER,
        ).export([analysis], tmp_path)

        combined = next(tmp_path.rglob("*_Complete_Packet.pdf"))
        resume = next(g for g in analysis.groups if g.document_type == "Resume")
        with open_pdf(combined) as document:
            assert document.page_count == len(resume.export_page_indexes)

    def test_a_candidate_with_none_of_the_chosen_types_is_skipped(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService(document_types=["Transcript"]).export([analysis], tmp_path)

        assert result.document_count == 0
        assert result.packet_count == 0, "an empty packet PDF was written"

    def test_analysis_is_untouched_by_the_filter(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        """Narrowing the output must not hide what was found."""
        analysis = pipeline.analyze_file(packet)
        before = [g.document_type for g in analysis.groups]
        ExportService(document_types=["Resume"]).export([analysis], tmp_path)
        assert [g.document_type for g in analysis.groups] == before


class TestRedundantPacketsAreSkipped:
    """A "complete packet" of one document is a duplicate, not a packet."""

    def test_no_packet_when_it_would_duplicate_a_single_document(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        ExportService(document_types=["Resume"]).export([analysis], tmp_path)

        written = sorted(p.name for p in tmp_path.rglob("*.pdf"))
        assert written == ["Benjamin_Perez_Resume.pdf"], written

    def test_a_packet_is_still_written_when_it_combines_documents(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        ExportService(document_types=["Resume", "Cover Letter"]).export(
            [analysis], tmp_path
        )

        written = sorted(p.name for p in tmp_path.rglob("*.pdf"))
        assert "Benjamin_Perez_Complete_Packet.pdf" in written

    def test_a_lone_document_still_gets_a_packet_when_that_is_the_only_output(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        """With separate documents off, the packet is the only thing asked for."""
        analysis = pipeline.analyze_file(packet)
        ExportService(
            document_types=["Resume"], export_separate_documents=False
        ).export([analysis], tmp_path)

        written = sorted(p.name for p in tmp_path.rglob("*.pdf"))
        assert written == ["Benjamin_Perez_Complete_Packet.pdf"], written


class TestBatchRunFolder:
    """One Sort & Save run, one timestamped folder.

    Repeated exports into the same chosen directory used to interleave, and
    afterwards nothing said which resume came from which run. The folder is
    opt-in (``batch_folder=True``) because a caller that names an exact
    destination -- a benchmark, most of the tests above -- means that
    destination, not a folder underneath it.
    """

    def run_folders(self, base: Path) -> list[Path]:
        return sorted(p for p in base.iterdir() if p.is_dir())

    def test_a_run_folder_is_created(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService(batch_folder=True).export([analysis], tmp_path)

        folders = self.run_folders(tmp_path)
        assert len(folders) == 1, folders
        assert result.output_directory == folders[0]
        assert result.output_directory.parent == tmp_path

    def test_the_name_is_a_windows_safe_timestamp(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService(batch_folder=True).export([analysis], tmp_path)

        name = result.output_directory.name
        assert re.fullmatch(r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-(AM|PM)", name), name
        assert not set(name) & set(':*?"<>|/\\')

    def test_naming_covers_midnight_noon_and_afternoon(self) -> None:
        assert batch_folder_name(datetime(2026, 8, 26, 10, 32)) == "2026-08-26_10-32-AM"
        assert batch_folder_name(datetime(2026, 8, 26, 0, 5)) == "2026-08-26_12-05-AM"
        assert batch_folder_name(datetime(2026, 8, 26, 12, 0)) == "2026-08-26_12-00-PM"
        assert batch_folder_name(datetime(2026, 8, 26, 13, 45)) == "2026-08-26_01-45-PM"

    def test_the_type_folders_live_inside_it(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService(
            batch_folder=True, group_by_document_type=True, export_combined_packets=False
        ).export([analysis], tmp_path)

        run = result.output_directory
        assert (run / "Resumes").is_dir()
        assert (run / "Cover Letters").is_dir()
        # ...and nothing was scattered at the base alongside the run folder.
        assert not list(tmp_path.glob("*.pdf"))
        assert [p.name for p in tmp_path.iterdir()] == [run.name]

    def test_two_runs_get_different_folders(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        service = ExportService(batch_folder=True)
        first = service.export([analysis], tmp_path)
        second = service.export([pipeline.analyze_file(packet)], tmp_path)

        assert first.output_directory != second.output_directory
        assert len(self.run_folders(tmp_path)) == 2

    def test_a_same_minute_collision_is_given_its_own_folder(
        self, tmp_path: Path
    ) -> None:
        """Two runs inside one minute share a name; they must not share a folder."""
        moment = datetime(2026, 8, 26, 10, 32)
        made = [create_batch_directory(tmp_path, moment=moment) for _ in range(3)]

        assert [p.name for p in made] == [
            "2026-08-26_10-32-AM",
            "2026-08-26_10-32-AM (2)",
            "2026-08-26_10-32-AM (3)",
        ]
        assert len({p for p in made}) == 3
        assert all(p.is_dir() for p in made)

    def test_a_folder_left_by_something_else_is_not_reused(
        self, tmp_path: Path
    ) -> None:
        moment = datetime(2026, 8, 26, 10, 32)
        (tmp_path / "2026-08-26_10-32-AM").mkdir()
        (tmp_path / "2026-08-26_10-32-AM" / "stale.pdf").write_bytes(b"%PDF-1.4\n")

        created = create_batch_directory(tmp_path, moment=moment)
        assert created.name == "2026-08-26_10-32-AM (2)"
        assert not list(created.iterdir()), "the run started in somebody else's folder"

    def test_every_pdf_in_one_batch_shares_one_folder(
        self, pipeline, samples_dir: Path, tmp_path: Path
    ) -> None:
        """Twenty inputs, one run folder -- not one folder per source PDF."""
        sources = [
            samples_dir / sample_data.sample_a().filename,
            samples_dir / sample_data.sample_b().filename,
            samples_dir / sample_data.sample_g().filename,
        ]
        analyses = [pipeline.analyze_file(path) for path in sources]
        result = ExportService(batch_folder=True).export(analyses, tmp_path)

        assert len(self.run_folders(tmp_path)) == 1
        run = result.output_directory
        assert result.document_count > len(sources)
        for document in result.exported:
            assert run in document.output_path.parents, document.output_path

    def test_a_resume_only_export_still_gets_its_run_folder(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        analysis = pipeline.analyze_file(packet)
        result = ExportService(
            batch_folder=True,
            group_by_document_type=True,
            export_combined_packets=False,
            document_types=["Resume"],
        ).export([analysis], tmp_path)

        run = result.output_directory
        assert run.parent == tmp_path
        assert (run / "Resumes").is_dir()
        assert not (run / "Cover Letters").exists()
        assert [d.group.document_type for d in result.exported] == ["Resume"]

    def test_filename_collisions_are_still_handled_inside_the_run(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        """Two identical candidates in one batch: the run folder must not
        change the ``_2`` suffix behaviour that stops an overwrite."""
        analyses = [pipeline.analyze_file(packet), pipeline.analyze_file(packet)]
        result = ExportService(
            batch_folder=True, group_by_document_type=True, export_combined_packets=False
        ).export(analyses, tmp_path)

        written = sorted(p.name for p in (result.output_directory / "Resumes").glob("*.pdf"))
        assert len(written) == 2, written
        assert len(set(written)) == 2, "one export overwrote the other"

    def test_nothing_to_export_leaves_no_empty_run_folder(
        self, tmp_path: Path
    ) -> None:
        """A misleading empty folder is worse than none at all."""
        result = ExportService(batch_folder=True).export([], tmp_path)

        assert result.document_count == 0
        assert self.run_folders(tmp_path) == []
        assert result.output_directory == tmp_path

    def test_an_unusable_base_directory_creates_nothing(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        """Validation failing before export begins must not leave a run folder."""
        blocked = tmp_path / "blocked"
        blocked.write_text("not a directory", encoding="utf-8")

        analysis = pipeline.analyze_file(packet)
        result = ExportService(batch_folder=True).export([analysis], blocked)

        assert result.has_errors
        assert result.document_count == 0
        assert blocked.is_file(), "the blocking file was replaced"

    def test_it_stays_off_unless_asked_for(
        self, pipeline, packet: Path, tmp_path: Path
    ) -> None:
        """The default writes exactly where the caller pointed."""
        analysis = pipeline.analyze_file(packet)
        result = ExportService().export([analysis], tmp_path)
        assert result.output_directory == tmp_path
