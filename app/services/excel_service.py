"""Excel index generation (``DocumentIndex.xlsx``) using openpyxl."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.packet import CandidatePacket
from app.services.confidence import ConfidenceThresholds
from app.services.export_service import ExportedDocument
from app.utils.logging_setup import get_logger

logger = get_logger("excel")

INDEX_FILENAME = "DocumentIndex.xlsx"

_COLUMNS: tuple[tuple[str, int], ...] = (
    ("Candidate Packet", 16),
    ("Candidate", 24),
    ("Email", 30),
    ("Phone", 18),
    ("LinkedIn", 32),
    ("Job", 28),
    ("Applicant ID", 14),
    ("Candidate Confidence", 20),
    ("Document Type", 20),
    ("Source PDF", 34),
    ("Source Pages", 14),
    ("Output File", 38),
    ("Combined Packet File", 38),
    ("Classification Confidence", 22),
    ("Boundary Confidence", 20),
    ("Review Status", 18),
    ("Processing Timestamp", 22),
)

#: 1-based indexes of the columns holding a 0..1 confidence, found by name so
#: inserting a column cannot leave the percentage formatting pointing elsewhere.
_CONFIDENCE_COLUMNS: tuple[int, ...] = tuple(
    index
    for index, (title, _width) in enumerate(_COLUMNS, start=1)
    if title.endswith("Confidence")
)

_HEADER_FILL = PatternFill("solid", start_color="FF1F3864", end_color="FF1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)


def _page_range(page_numbers: Sequence[int]) -> str:
    if not page_numbers:
        return ""
    ordered = sorted(page_numbers)
    if len(ordered) == 1:
        return str(ordered[0])
    if ordered == list(range(ordered[0], ordered[-1] + 1)):
        return f"{ordered[0]}-{ordered[-1]}"
    return ", ".join(str(number) for number in ordered)


def _packet_label(packet: CandidatePacket | None) -> str:
    """A short, sortable packet identifier for the spreadsheet."""
    if packet is None:
        return ""
    if packet.is_unknown:
        return "Unassigned"
    return packet.id


def _review_status(document: ExportedDocument, thresholds: ConfidenceThresholds) -> str:
    """The worst of the three questions asked about this document."""
    group = document.group
    if group.type_manually_set and group.association_manually_set:
        return "Corrected by reviewer"
    if group.association_review and not group.association_manually_set:
        return "Candidate needs review"
    if group.type_manually_set:
        return "Corrected by reviewer"
    return thresholds.band(group.overall_confidence).label


def write_document_index(
    documents: Sequence[ExportedDocument],
    output_directory: str | Path,
    *,
    thresholds: ConfidenceThresholds | None = None,
    timestamp: datetime | None = None,
    filename: str = INDEX_FILENAME,
    packets: Sequence[CandidatePacket] = (),
) -> Path | None:
    """Write a professional, filterable index of everything exported.

    ``packets`` lets each row name the candidate it was filed under, which is
    not always the name written on the document itself -- an anonymous cover
    letter attributed by context has an owner but no name of its own.

    Returns the workbook path, or ``None`` when there was nothing to write or
    the file could not be saved (never raises into the export flow).
    """
    if not documents:
        return None

    thresholds = thresholds or ConfidenceThresholds()
    stamp = timestamp or datetime.now()
    destination = Path(output_directory) / filename

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Document Index"

    sheet.append([title for title, _width in _COLUMNS])
    for index, (_title, width) in enumerate(_COLUMNS, start=1):
        column = get_column_letter(index)
        sheet.column_dimensions[column].width = width
        cell = sheet.cell(row=1, column=index)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 30

    packets_by_id = {packet.id: packet for packet in packets}

    for document in documents:
        group = document.group
        packet = packets_by_id.get(group.packet_id)
        candidate = packet.candidate if packet and packet.candidate.name else group.candidate
        sheet.append(
            [
                _packet_label(packet),
                candidate.name or "Unknown",
                candidate.email or "",
                candidate.phone or "",
                candidate.linkedin or "",
                candidate.job_title or "",
                candidate.applicant_id or "",
                round(group.association_confidence, 3),
                group.document_type,
                document.source_pdf.name,
                _page_range(document.page_numbers),
                document.output_path.name,
                Path(group.packet_export_path).name if group.packet_export_path else "",
                round(group.classification_confidence, 3),
                round(group.boundary_confidence, 3),
                _review_status(document, thresholds),
                stamp.strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

    last_row = sheet.max_row
    last_column = get_column_letter(len(_COLUMNS))

    # Freeze the header, enable filtering, and show confidences as percentages.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{last_column}{last_row}"
    # Derived from the headers rather than hardcoded: adding a column used to
    # silently format the wrong ones as percentages.
    for column_index in _CONFIDENCE_COLUMNS:
        for row in sheet.iter_rows(
            min_row=2, min_col=column_index, max_col=column_index, max_row=last_row
        ):
            for cell in row:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="center")

    try:
        workbook.save(destination)
    except OSError as exc:
        logger.warning("Could not write the Excel index: %s", exc)
        return None
    finally:
        workbook.close()

    logger.info("Wrote Excel index with %s rows", last_row - 1)
    return destination


__all__ = ["write_document_index", "INDEX_FILENAME"]
